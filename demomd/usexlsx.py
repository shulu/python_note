#!/usr/bin/env python3
# -*- coding: utf-8 -*-

__author__ = 'SarcasMe'

import pandas as pd
import numpy as np
import openpyxl

#print(np.random.randint(low=0, high=10, size=(5, 5)))

#writer = pd.ExcelWriter('pipelines.xlsx')
#df = pd.DataFrame([[1,2,3,4,5,6,7]], columns=['pipeline_id', 'author', 'author_email', 'commit_id', 'commit_message', 'branch', 'pipeline_status'])
#df = pd.DataFrame([[1,2,3,4,5,6,7]])
#df.append(df)
#print(df)
#df.to_excel(writer,'Sheet1')
#writer.save()

wbname = 'pipelines.xlsx'
wb = openpyxl.load_workbook(wbname)

sheet = wb.active
# print(wb.sheetnames)
##sheet = wb["Sheet1"]
row = sheet.max_row # <-最大行数
max_row = row+1
# for row1 in range(2,len(data)+2):  # 写入数据
#    for col1 in range(1,len(data[row1-2])+1):
datas = [9, 8, 7, 6, 5, 4, 3]
for col in range(1, len(datas)+1):
    # print(max_row, col1)
    _ = sheet.cell(row=max_row, column=col, value=str(datas[col-1]))

wb.save(filename=wbname)
print("保存成功")