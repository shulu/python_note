#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import requests
import json
import math
import time
import re
from openpyxl import load_workbook
import pandas as pd

__author__ = 'SarcasMe'

class my_pipelines():

    def __init__(self):

        self.project = 'banggood/www'
        self.private_token = 's64NyxXzb-Dreu9-4iM3'
        self.excel_name = 'pipelines_bak.xlsx'
        self.max_id = 0
        self.max_count = 0

    def continue_to_excel(self):

        """
        获取最新数据计算是否需要更新本地数据
        :return:
        """
        data = open('max_count.json').read()
        data = json.loads(data)
        last_allacount = data['all_count']
        data = self.return_pipelines(False)
        all_commit = data['count']['all']
        pages = 0
        if all_commit > last_allacount:
            df = pd.DataFrame(pd.read_excel(self.excel_name))
            self.max_id = df['pipeline_id'].tail(1).values[0]
            self.max_count = all_commit
            pages = int(math.ceil((all_commit-last_allacount) / 20))
        else:
            print('当前已经时最新数据')
        return pages

    def get_pages(self):

        data = self.return_pipelines(False)
        all_commit = data[ 'count' ][ 'all' ]
        pages = int(math.ceil(all_commit / 20))
        self.record_max(all_commit)
        return pages

    def get_pipelines(self, pages):

        """
        分离出数据返回存储
        :param pages:
        :return:
        """
        while pages > 0:

            time.sleep(0.5)
            print('start page '+str(pages))
            data = self.return_pipelines(pages)

            pipelines = data['pipelines']
            pipelines.reverse()
            for pipe in pipelines:

                branch = author_email = commit_id = commit_message = 'unknow'
                pipeline_id = pipe['id']
                if pipeline_id <= self.max_id:
                    continue
                author = pipe['user']['name']
                if pipe['commit']:
                    commit_id = pipe['commit']['id']
                    author_email = pipe['commit']['author_email']
                    commit_message = pipe['commit']['title']
                commit_date = self.conv_utc_time(pipe['created_at'])
                if pipe['ref']['branch']:
                    branch = pipe['ref']['name']
                pipeline_status = pipe['details']['status']['label']
                pipeline = [
                    pipeline_id, author, author_email, branch,
                    commit_id, commit_message, commit_date, pipeline_status
                ]
                yield pipeline
            pages -= 1

    @staticmethod
    def conv_utc_time(dt):

        """
        转换UTC时间为正常时间
        :param dt:
        :return:
        """
        dates = re.search("(\d*-\d*-\d*)", dt).group()
        times = re.search("(\d*:\d*:\d*)", dt).group()
        dt = dates + ' ' + times
        # 转换成时间数组
        timeArray = time.strptime(dt, "%Y-%m-%d %H:%M:%S")
        # 转换成时间戳
        timestamp = time.mktime(timeArray)
        timestamp += 8 * 3600
        # 转换成localtime
        time_local = time.localtime(timestamp)
        # 转换成新的时间格式(2016-05-05 20:28:54)
        dt = time.strftime("%Y-%m-%d %H:%M:%S", time_local)
        return dt

    def return_pipelines(self, page):

        """
        返回json数据
        :param page:
        :return:
        """
        page_ = 1
        if page:
            page_ = page
        url = 'https://g.banggood.com/{}/pipelines.json?scope=all&page={}&private_token={}'.format(self.project, page_, self.private_token)
        ret = requests.get(url)
        if ret.status_code == 200:
            data = json.loads(ret.text)
            return data
        else:
            return False

    def save_to_excel(self):

        """
        初始化存储数据
        :return:
        """
        pages = self.continue_to_excel()
        wb = load_workbook(self.excel_name)
        sheet = wb.active
        row = sheet.max_row  # <-最大行数
        max_row = row + 1
        for pipeline in self.get_pipelines(pages=pages):
            for col in range(1, len(pipeline) + 1):
                _ = sheet.cell(row=max_row, column=col, value=str(pipeline[col - 1]))
            max_row += 1
        wb.save(filename=self.excel_name)
        self.record_max(self.max_count)
        print("保存成功")

    @staticmethod
    def record_max(count):
        with open("max_count.json", "w") as f:
            json.dump({'all_count':count}, f)
        print("加载count文件完成...")

if __name__ == '__main__':

    pipe = my_pipelines()
    pipe.save_to_excel()
    #pipe.continue_to_excel()